# from openpyxl import Workbook
# from openpyxl.styles import Font

# wb = Workbook()
# sheet = wb.active
# sheet.title = "Language"
# wb.create_sheet(title="Capital")

# lang = ["Kannada", "Telugu", "Tamil"]
# state = ["Karnataka", "Telangana", "Tamil Nadu"]
# capital = ["Bengaluru", "Hyderabad", "Chennai"]
# code = ['KA', 'TS', 'TN']

# sheet.cell(row=1, column=1).value = "State"
# sheet.cell(row=1, column=2).value = "Language"
# sheet.cell(row=1, column=3).value = "Code"

# ft = Font(bold=True)

# for row in sheet["A1:C1"]:
#     for cell in row:
#         cell.font = ft

# for i in range(2, 5):
#     sheet.cell(row=i, column=1).value = state[i-2]
#     sheet.cell(row=i, column=2).value = lang[i-2]
#     sheet.cell(row=i, column=3).value = code[i-2]

# wb.save("demo.xlsx")

# sheet = wb["Capital"]
# sheet.cell(row=1, column=1).value = "State"
# sheet.cell(row=1, column=2).value = "Capital"
# sheet.cell(row=1, column=3).value = "Code"

# for row in sheet["A1:C1"]:
#     for cell in row:
#         cell.font = ft

# for i in range(2, 5):
#     sheet.cell(row=i, column=1).value = state[i-2]
#     sheet.cell(row=i, column=2).value = capital[i-2]
#     sheet.cell(row=i, column=3).value = code[i-2]

# wb.save("demo.xlsx")

# sheet = wb["Language"]
# srchCode = input("Enter state code for finding language: ")

# for i in range(2, 5):
#     data = sheet.cell(row=i, column=3).value
#     if data == srchCode:
#         print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# sheet = wb["Capital"]
# srchCode = input("Enter state code for finding capital: ")

# for i in range(2, 5):
#     data = sheet.cell(row=i, column=3).value
#     if data == srchCode:
#         print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# wb.close()

# import Section
from openpyxl import Workbook
from openpyxl.styles import Font

# Defining Variables, lists and data
wb = Workbook()
sheets = {"Language": [], "Capital": []}

data = {
    "Language": [("Kannada", "Karnataka", "KA"),
                  ("Telugu", "Telangana", "TS"), 
                  ("Tamil", "Tamil Nadu", "TN")],
    "Capital": [("Bengaluru", "Karnataka", "KA"), 
                ("Hyderabad", "Telangana", "TS"), 
                ("Chennai", "Tamil Nadu", "TN")]
}
titles = {
    "Language": ["Language", "State", "Code"],
    "Capital": ["Capital", "State", "Code"]
}

# Creating a spreadSheet of above data
for sheet_name, info in data.items():
    sheet = wb.create_sheet(title=sheet_name)
    sheets[sheet_name] = info
    for row in info:
        sheet.append(row)
    for col_idx, title in enumerate(titles[sheet_name]):
        cell = sheet.cell(row=1, column=col_idx + 1)
        cell.value = title
        cell.font = Font(bold=True)
wb.save("demo.xlsx")

# Searching and printing the output
for sheet_name, search_type in [("Language", "language"), ("Capital", "capital")]:
    sheet = wb[sheet_name]
    srchCode = input(f"Enter state code for finding {search_type}: ")
    for row in sheets[sheet_name]:
        if row[2] == srchCode:
            print(f"Corresponding {search_type} for code {srchCode} is {row[0]}")
wb.close()

