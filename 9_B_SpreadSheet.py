# Demonstrate python program to read the data from the spreadsheet and write the data in to the spreadsheet

# import Section
from openpyxl import Workbook
from openpyxl.styles import Font

# Defining Variables, lists and data
wb = Workbook()
sheets = {"Language": [], "Capital": []}

data = {
    "Language": [("Kannada", "Karnataka", "KA"),
                  ("Telugu", "AndhraPradesh", "AP"), 
                  ("Tamil", "Tamil Nadu", "TN")],
    "Capital": [("Bengaluru", "Karnataka", "KA"), 
                ("Hyderabad", "AndhraPradesh", "AP"), 
                ("Chennai", "Tamil Nadu", "TN")]
}
titles = {
    "Language": ["Language", "State", "Code"],
    "Capital": ["Capital", "State", "Code"]
}

# Creating a spreadSheet of above data
for sheet_name, info in data.items():
    sheet = wb.create_sheet(title=sheet_name)
    sheets[sheet_name] = info
    for row in info:
        sheet.append(row)
    for col_idx, title in enumerate(titles[sheet_name]):
        cell = sheet.cell(row=1, column=col_idx + 1)
        cell.value = title
        cell.font = Font(bold=True)
wb.save("stateInfo.xlsx")

# Searching and printing the output
for sheet_name, search_type in [("Language", "language"), ("Capital", "capital")]:
    sheet = wb[sheet_name]
    srchCode = input(f"Enter state code for finding {search_type}: ")
    for row in sheets[sheet_name]:
        if row[2] == srchCode:
            print(f"Corresponding {search_type} for code {srchCode} is {row[0]}")
wb.close()